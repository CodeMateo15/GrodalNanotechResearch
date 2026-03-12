"""
Convert txt files into Excel format matching Combined_community_worksheet.xlsx.

HOW TO USE:
1. Place this script in the same folder as your .txt files
2. Edit the FILE_SOURCE_MAP below to match your filenames, source names, numbers, and title style
3. Run the script in VS Code (press F5 or click Run)

The output file will be saved in the same folder as this script.
"""

import re
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIGURATION ────────────────────────────────────────────────────────────

# Map each txt filename to:
#   (source_name, source_number, title_style)
#
# title_style options:
#   "first_line"     — use the first non-blank line of the chunk as the title
#                      (good for Government, Newspapers, etc.)
#   "after_label"    — use the line after a section label like "Reports" or
#                      "Policy Forum" as the title
#                      (good for Science Research, Science News)
#   "business_press" — use sequential "Article N" as the title; body starts
#                      after the Copyright line
#                      (good for Business Press, Business Week, WSJ, etc.)
#
FILE_SOURCE_MAP = {
    "government.txt":       ("Government",       1, "first_line"),
    "Science_research.txt": ("Science Research", 3, "after_label"),
    "business_press.txt":   ("Business Press",   2, "business_press"),
    # "news_test.txt":      ("Newspapers",       2, "first_line"),
    # "scin_test.txt":      ("Science News",     4, "after_label"),
}

# Only keep articles dated within this year range.
# Articles with NO date found are always kept.
YEAR_MIN = 1983
YEAR_MAX = 2005

# Name of the output Excel file (saved in the same folder as this script)
OUTPUT_FILE = "output.xlsx"

# ── END CONFIGURATION ─────────────────────────────────────────────────────────

# Separator must be 5+ asterisks that are the ONLY non-whitespace content on
# their line — prevents stray * in bullet points or footnotes from splitting.
SEPARATOR = re.compile(r'^\s*\*{5,}\s*$', re.MULTILINE)

# Section labels that precede the title in Science-style files
SECTION_LABELS = re.compile(
    r'^(Reports?|Policy\s+Forum|News|Perspective|Review|Letter|Editorial|'
    r'Research\s+Article|Brief\s+Communication)$',
    re.IGNORECASE
)

# Lines that signal the start of a references/notes section to strip
REFERENCES_START = re.compile(
    r'^(References(\s+and\s+Notes)?|Bibliography|Notes|Supporting\s+(Online\s+)?Material|'
    r'SOM\s+Text|Acknowledgements?|Supplementary)$',
    re.IGNORECASE
)

# Copyright line marks end of metadata in business press articles
COPYRIGHT_LINE = re.compile(r'^\(?\s*Copyright', re.IGNORECASE)

DATE_PATTERNS = [
    (r'\b(\d{1,2}\s+\w+\s+\d{4})\b',   '%d %B %Y'),
    (r'\b(\w+\s+\d{1,2},\s+\d{4})\b',  '%B %d, %Y'),
    (r'\b(\d{1,2}\s+\w{3}\s+\d{4})\b', '%d %b %Y'),
    (r'\b(\w{3}\s+\d{1,2},\s+\d{4})\b','%b %d, %Y'),
    (r'\b(\d{4}-\d{2}-\d{2})\b',        '%Y-%m-%d'),
]

MONTH_NAMES = {
    'january','february','march','april','may','june',
    'july','august','september','october','november','december',
    'jan','feb','mar','apr','jun','jul','aug','sep','oct','nov','dec'
}

# Keywords for each topic column — add/remove terms as needed
#
# Each value is a list of Python regular-expression patterns used to
# search the article body. Notes for future edits:
# - Patterns are applied with `re.IGNORECASE` so no need for case variants.
# - Use `\b` to anchor word boundaries. Example:
#     r'\bnano\b'  -> matches the whole word "nano" only (not "nanotech").
# - Use `\w*` to match word characters after a stem. Example:
#     r'\bnano\w*' -> matches "nano", "nanotech", "nano123",
#     "nano_scale" but NOT "nano-scale" (hyphen is not a word character).
# - To include hyphens or dots, extend the character class:
#     r'\bnano[\w.-]*'  to allow "nano-scale" or "nano.tech".
#
TOPIC_KEYWORDS = {
    "Space":            [r'\bspace\b', r'\bsatellite', r'\baerospace', r'\borbit', r'\brocket', r'\bnasa\b', r'\bspacecraft'],
    "Electronics":      [r'\belectronic', r'\bsemiconductor', r'\bcircuit', r'\btransistor', r'\bdiode', r'\bchip\b'],
    "Artificial Intelligence": [r'\bartificial intelligence\b', r'\bmachine learning\b', r'\bneural network', r'\bdeep learning\b', r'\bai\b', r'\bgradient descent\b'],
    "Photonics":        [r'\bphotonic', r'\boptical\b', r'\blaser', r'\bfiber optic', r'\bphoton'],
    "Biotech/Biology":  [r'\bbiotech', r'\bbiology\b', r'\bbiological\b', r'\bgenetic', r'\bgenome', r'\bprotein\b', r'\bcell\b', r'\bbacterial'],
    "Semiconductors":   [r'\bsemiconductor', r'\bsilicon\b', r'\bgallium', r'\bdoping\b', r'\bwafer'],
    "Robotics":         [r'\brobot', r'\bautonomo', r'\bmanipulator', r'\bactuat'],
    "Computers/Computing": [r'\bcomput', r'\bprocessor', r'\bsoftware\b', r'\bhardware\b', r'\bwireless\b', r'\bdigital\b', r'\bmicroprocessor'],
    "Material Science": [r'\bmaterial science', r'\bcomposite\b', r'\balloy\b', r'\bpolymer\b', r'\bceramics?\b', r'\bcoating'],
    "Cleantech":        [r'\bcleantech\b', r'\brenewable', r'\bsolar\b', r'\bwind energy', r'\bclean energy', r'\bgreen tech', r'\bhydrogen fuel'],
    "Hypertext":        [r'\bhypertext\b', r'\bhyperlink', r'\bhtml\b', r'\bweb page', r'\bwww\b'],
    "Internet":         [r'\binternet\b', r'\bonline\b', r'\bworld wide web\b', r'\bbroadband', r'\bnetwork\b'],
    "Chemistry":        [r'\bchemi', r'\bmolecul', r'\breaction\b', r'\bcatalys', r'\bcompound\b', r'\bsynthes'],
    "Physics":          [r'\bphysics\b', r'\bphysical\b', r'\bquantum\b', r'\bthermodynamic', r'\bmechanics\b', r'\belectromagnet'],
    # "Nanotech" should capture variants like "nanotechnology".
    "Nanotech":         [r'\bnanotech\w*'],
    # "Nano" uses \w* to capture "nano", "nanotech", "nano123", etc.
    "Nano":             [r'\bnano\w*'],
}


def extract_date(text):
    for pattern, fmt in DATE_PATTERNS:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            candidate = match.group(1)
            parts = re.split(r'[\s,]+', candidate)
            has_month = any(p.lower() in MONTH_NAMES for p in parts)
            if not has_month and fmt != '%Y-%m-%d':
                continue
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def count_keyword(text, patterns):
    return sum(len(re.findall(p, text, re.IGNORECASE)) for p in patterns)


def count_words(text):
    return len(text.split())


def get_non_blank_lines(text):
    return [l.strip() for l in text.splitlines() if l.strip()]


def strip_references(text):
    """Split text into (body, references) at the first references heading."""
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if REFERENCES_START.match(line.strip()):
            return "\n".join(lines[:i]).strip(), "\n".join(lines[i:]).strip()
    return text, ""


def extract_title(chunk, title_style, article_num=None):
    lines = get_non_blank_lines(chunk)
    if not lines:
        return ""

    if title_style == "first_line":
        return lines[0]

    elif title_style == "after_label":
        for i, line in enumerate(lines):
            if SECTION_LABELS.match(line) and i + 1 < len(lines):
                return lines[i + 1]
        return lines[0]

    elif title_style == "business_press":
        return f"Article {article_num}"

    return lines[0]


def extract_body(chunk, title_style):
    """
    Return (body, references): body is clean article text after the title,
    references is everything from the references heading onward.
    """
    lines = chunk.splitlines()
    non_blank = [(i, l.strip()) for i, l in enumerate(lines) if l.strip()]
    if not non_blank:
        return "", ""

    if title_style == "first_line":
        title_idx = non_blank[0][0]
        raw = "\n".join(lines[title_idx + 1:]).strip()

    elif title_style == "after_label":
        raw = chunk  # fallback
        for i, (line_idx, line) in enumerate(non_blank):
            if SECTION_LABELS.match(line) and i + 1 < len(non_blank):
                title_line_idx = non_blank[i + 1][0]
                raw = "\n".join(lines[title_line_idx + 1:]).strip()
                break
        else:
            title_idx = non_blank[0][0]
            raw = "\n".join(lines[title_idx + 1:]).strip()

    elif title_style == "business_press":
        # Body starts on the line after the Copyright line
        raw = chunk  # fallback
        for i, line in enumerate(lines):
            if COPYRIGHT_LINE.match(line.strip()):
                raw = "\n".join(lines[i + 1:]).strip()
                break

    else:
        raw = chunk

    return strip_references(raw)


def in_year_range(year):
    if year is None:
        return True  # Keep articles with no date
    return YEAR_MIN <= year <= YEAR_MAX


def parse_articles(filepath, source_name, source_num, title_style):
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    content = content.replace('\r\n', '\n').replace('\r', '\n')
    chunks = [c.strip() for c in SEPARATOR.split(content) if c.strip()]

    rows = []
    skipped = 0
    article_num = 0  # counter for business_press sequential titles

    for chunk in chunks:
        date = extract_date(chunk)
        year = date.year if date else None

        if not in_year_range(year):
            skipped += 1
            continue

        article_num += 1
        title = extract_title(chunk, title_style, article_num=article_num)
        body, refs = extract_body(chunk, title_style)

        row = {
            'Date':       date.strftime('%d %B %Y') if date else None,
            'Year':       year,
            'Sources':    source_num,
            'Name':       source_name,
            'Word count': count_words(body),
            'Title':      title,
            'Body':       body,
            'References': refs,
        }
        for topic, patterns in TOPIC_KEYWORDS.items():
            row[topic] = count_keyword(body, patterns)

        rows.append(row)

    if skipped:
        print(f"  (skipped {skipped} articles outside {YEAR_MIN}–{YEAR_MAX})")

    return rows


def write_excel(all_rows, output_path):
    columns = (
        ['Date', 'Year', 'Sources', 'Name', 'Word count', 'Title', 'Body', 'References']
        + list(TOPIC_KEYWORDS.keys())
    )
    df = pd.DataFrame(all_rows, columns=columns)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        ws = writer.sheets['Sheet1']

        header_fill = PatternFill('solid', start_color='4472C4', end_color='4472C4')
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        fixed_widths = {'A': 14, 'B': 6, 'C': 8, 'D': 18, 'E': 12}
        all_cols = [chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26) for i in range(ws.max_column)]
        for col_idx, col_letter in enumerate(all_cols, start=1):
            col_name = df.columns[col_idx - 1]
            if col_name == 'Title':
                ws.column_dimensions[col_letter].width = 40
            elif col_name in ('Body', 'References'):
                ws.column_dimensions[col_letter].width = 60
            else:
                ws.column_dimensions[col_letter].width = fixed_widths.get(col_letter, 16)

        data_font = Font(name='Arial', size=11)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = Alignment(wrap_text=False, vertical='top')


def main():
    script_dir = Path(__file__).parent
    all_rows = []

    for filename, (source_name, source_num, title_style) in FILE_SOURCE_MAP.items():
        filepath = script_dir / filename
        if not filepath.exists():
            print(f"WARNING: '{filename}' not found in {script_dir} — skipping.")
            continue

        rows = parse_articles(filepath, source_name, source_num, title_style)
        all_rows.extend(rows)
        print(f"'{filename}' ({source_name}, Sources={source_num}): {len(rows)} articles kept")
        for i, r in enumerate(rows, 1):
            print(f"  Article {i}: Date={r['Date']}, Words={r['Word count']}, Title='{r['Title'][:60]}'")

    if not all_rows:
        print("No articles found. Check that your txt files are in the same folder as this script.")
        return

    output_path = script_dir / OUTPUT_FILE
    write_excel(all_rows, output_path)
    print(f"\nDone! {len(all_rows)} total rows written to '{output_path}'")


if __name__ == '__main__':
    main()