"""
Convert txt files into Excel format matching Combined_community_worksheet.xlsx.

HOW TO USE:
1. Place this script in the same folder as your .txt files
2. Edit the FILE_SOURCE_MAP below to match your filenames to their source names and numbers
3. Run the script in VS Code (press F5 or click Run)

The output file will be saved in the same folder as this script.
"""

import re
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIGURATION ────────────────────────────────────────────────────────────

# Map each txt filename to (source name, source number).
# Add or remove entries to match the files in your folder.
# Valid source names: Government, Science Research, Science News,
#                     Business Press, Business, Futurists, Newspapers
FILE_SOURCE_MAP = {
    "govt_test.txt":    ("Government",       1),
    "Sci_res_test.txt": ("Science Research", 3),
    # "news_test.txt":  ("Newspapers",       7),  # ← uncomment and edit to add more
}

# Name of the output Excel file (saved in the same folder as this script)
OUTPUT_FILE = "output.xlsx"

# ── END CONFIGURATION ─────────────────────────────────────────────────────────

SEPARATOR = re.compile(r'\*{5,}')

DATE_PATTERNS = [
    (r'\b(\d{1,2}\s+\w+\s+\d{4})\b',   '%d %B %Y'),
    (r'\b(\w+\s+\d{1,2},\s+\d{4})\b',  '%B %d, %Y'),
    (r'\b(\d{1,2}\s+\w{3}\s+\d{4})\b', '%d %b %Y'),
    (r'\b(\w{3}\s+\d{1,2},\s+\d{4})\b','%b %d, %Y'),
    (r'\b(\d{4}-\d{2}-\d{2})\b',        '%Y-%m-%d'),
]

MONTH_NAMES = {
    'january','february','march','april','may','june',
    'july','august','september','october','november','december',
    'jan','feb','mar','apr','jun','jul','aug','sep','oct','nov','dec'
}


def extract_date(text):
    for pattern, fmt in DATE_PATTERNS:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            candidate = match.group(1)
            parts = re.split(r'[\s,]+', candidate)
            has_month = any(p.lower() in MONTH_NAMES for p in parts)
            if not has_month and fmt != '%Y-%m-%d':
                continue
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def count_nanotech(text):
    return len(re.findall(r'\bnano\w*', text, re.IGNORECASE))


def count_words(text):
    return len(text.split())


def parse_articles(filepath, source_name, source_num):
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    content = content.replace('\r\n', '\n').replace('\r', '\n')
    chunks = [c.strip() for c in SEPARATOR.split(content) if c.strip()]

    rows = []
    for chunk in chunks:
        date = extract_date(chunk)
        rows.append({
            'Date':       date,
            'Year':       date.year if date else None,
            'Nanotech':   count_nanotech(chunk),
            'Sources':    source_num,
            'Name':       source_name,
            'Word count': count_words(chunk),
        })
    return rows


def write_excel(all_rows, output_path):
    df = pd.DataFrame(all_rows, columns=['Date', 'Year', 'Nanotech', 'Sources', 'Name', 'Word count'])

    with pd.ExcelWriter(output_path, engine='openpyxl', datetime_format='YYYY-MM-DD') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        ws = writer.sheets['Sheet1']

        header_fill = PatternFill('solid', start_color='4472C4', end_color='4472C4')
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for col, width in {'A': 14, 'B': 6, 'C': 10, 'D': 8, 'E': 18, 'F': 12}.items():
            ws.column_dimensions[col].width = width

        data_font = Font(name='Arial', size=11)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font


def main():
    script_dir = Path(__file__).parent
    all_rows = []

    for filename, (source_name, source_num) in FILE_SOURCE_MAP.items():
        filepath = script_dir / filename
        if not filepath.exists():
            print(f"WARNING: '{filename}' not found in {script_dir} — skipping.")
            continue

        rows = parse_articles(filepath, source_name, source_num)
        all_rows.extend(rows)
        print(f"'{filename}' ({source_name}, Sources={source_num}): {len(rows)} articles parsed")
        for i, r in enumerate(rows, 1):
            print(f"  Article {i}: Date={r['Date']}, Nanotech={r['Nanotech']}, Words={r['Word count']}")

    if not all_rows:
        print("No articles found. Check that your txt files are in the same folder as this script.")
        return

    output_path = script_dir / OUTPUT_FILE
    write_excel(all_rows, output_path)
    print(f"\nDone! {len(all_rows)} total rows written to '{output_path}'")


if __name__ == '__main__':
    main()