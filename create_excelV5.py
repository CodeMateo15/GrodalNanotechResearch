"""
Convert txt files into Excel format matching Combined_community_worksheet.xlsx.

V5: Adds Newspaper source, replaces Business RTF with Business TXT,
    fixes Science separator threshold, adds Molecular Engineering /
    Financial / Commerce / Application keywords, dual date columns
    with match percentage.

HOW TO USE:
1. Place this script in the same folder as your .txt files
2. Run: python create_excelV5.py
"""

import re
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIGURATION ────────────────────────────────────────────────────────────

FILE_SOURCE_MAP = {
    "government.txt":           ("Government",        1, "government"),
    "Science_news.txt":         ("Science News",       2, "after_label"),
    "Science_research.txt":     ("Science Research",   3, "after_label"),
    "Business_press.txt":       ("Business Press",     4, "business_press"),
    "Business.txt":             ("Business",           5, "business_press"),
    "Old Python Files/Business_2005.rtf": ("Business", 5, "business_press"),
    "futurists.txt":            ("Futurists",          6, "futurist"),
    "Newspapers_1984-2005.txt": ("Newspapers",         7, "newspaper"),
}

YEAR_MIN = 1983
YEAR_MAX = 2005

OUTPUT_FILE = "output.xlsx"

# Reference worksheet with dates for sources that lack them in the raw text
REFERENCE_WORKSHEET = "Combined community worksheet.xlsx"

# ── SEPARATORS ───────────────────────────────────────────────────────────────

SEP_ASTERISKS = re.compile(r'^\s*\*{19,}\s*$', re.MULTILINE)
SEP_BUSINESS  = re.compile(r'^\s*Article\s+\d+\s+\*{19,}', re.MULTILINE)
SEP_NEWSPAPER = re.compile(
    # Factiva-style boundaries. Observed variants in Newspapers_1984-2005.txt:
    #   "Article N ***..."                        3361 lines
    #   "Article N Previous Article|***..."         87 lines (pipe separator)
    #   "Article N Previous Article ***..."          8 lines (space separator)
    #   "Article N Next Article"                     5 lines (no asterisks)
    #   "Article N" bare                             2 lines
    r'(?:'
    r'^\s*Article\s+\d+'
    r'(?:(?:\s+(?:Previous|Next)\s+Article)?\s*\|?\s*\*{10,})?'
    r'\s*$'
    r'|^\s*\*\s+\*\s+\*\s*$'
    r')',
    re.MULTILINE | re.IGNORECASE,
)

# ── SECTION LABELS (Science-style) ──────────────────────────────────────────

SECTION_LABELS = re.compile(
    r'^(Reports?|Policy\s+Forum|News|Perspectives?|Reviews?|Letters?|'
    r'Editorial|Research\s+Article|Brief\s+Communications?|Brevia|'
    r'Random\s+Samples?|Newsmakers?|News\s+Focus|'
    r'News\s+of\s+the\s+Week|Findings|'
    r'This\s+Week\s+in\s+Science|Essays?|Corrections?)$',
    re.IGNORECASE
)

# ── REFERENCES ───────────────────────────────────────────────────────────────

REFERENCES_START = re.compile(
    r'^(References(\s+and\s+Notes)?|Bibliography|Notes|'
    r'Supporting\s+(Online\s+)?Material|SOM\s+Text|'
    r'Acknowledgements?|Supplementary\s+Materials?)$',
    re.IGNORECASE
)

# ── BUSINESS PRESS ───────────────────────────────────────────────────────────

COPYRIGHT_LINE = re.compile(r'^\(?\s*Copyright|\(c\)\s*\d{4}', re.IGNORECASE)

BUSINESS_TAIL_JUNK = re.compile(
    r'^\s*(Document\s+\S+|More Like This|Retrieving article\(s\)\.\.\.'
    r'|(?:Article\s+\d+\s+)?Next Article)\s*$',
    re.IGNORECASE
)

BUSINESS_HEADER_JUNK = re.compile(
    r'^\s*(Retrieving article|(?:Article\s+\d+\s+)?Next Article|'
    r'D&B Report|Financial Snapshot|Company Quick Search|'
    r'Company Data from|Add to Company List|'
    r'Analyst Report|Comparison Report|Quote|News|Details|Chart|'
    r'There is no related information)',
    re.IGNORECASE
)

# ── SCIENCE METADATA ────────────────────────────────────────────────────────

SCIENCE_META = re.compile(
    r'^(\d{1,5}\s*$|'
    r'DOI:\s|doi:\s|'
    r'Vol\.\s*\d|'
    r'Prev\s*\|\s*Table of Contents|'
    r'Science\s+\d{1,2}\s+\w+\s+\d{4}|'
    r'Science,\s+New\s+Series|'
    r'Copyright\s*\(?|'
    r'Originally\s+published|'
    r'\d{1,2}\s+\w+\s+\d{4}\s*$|'
    r'\d{1,2}\s+\w+\s+\d{4}\s+VOL\s+\d|'
    r'\*?\s*To whom correspondence)',
    re.IGNORECASE
)

# ── DATES ────────────────────────────────────────────────────────────────────

DATE_PATTERNS = [
    (r'\b(\d{1,2}\s+\w+\s+\d{4})\b',    '%d %B %Y'),
    (r'\b(\w+\s+\d{1,2},\s+\d{4})\b',   '%B %d, %Y'),
    (r'\b(\d{1,2}\s+\w{3}\s+\d{4})\b',  '%d %b %Y'),
    (r'\b(\w{3}\s+\d{1,2},\s+\d{4})\b', '%b %d, %Y'),
    (r'\b(\d{4}-\d{2}-\d{2})\b',         '%Y-%m-%d'),
]

MONTH_NAMES = {
    'january','february','march','april','may','june',
    'july','august','september','october','november','december',
    'jan','feb','mar','apr','jun','jul','aug','sep','oct','nov','dec'
}

# ── TOPIC KEYWORDS ───────────────────────────────────────────────────────────

TOPIC_KEYWORDS = {
    "Space":            [r'\bspace\b', r'\bsatellite', r'\baerospace', r'\borbit', r'\brocket', r'\bnasa\b', r'\bspacecraft'],
    "Electronics":      [r'\belectronic', r'\bcircuit', r'\btransistor', r'\bdiode', r'\bchip\b'],
    "Artificial Intelligence": [r'\bartificial intelligence\b', r'\bmachine learning\b', r'\bneural network', r'\bdeep learning\b', r'\bai\b', r'\bgradient descent\b'],
    "Photonics":        [r'\bphotonic', r'\boptical\b', r'\blaser', r'\bfiber optic', r'\bphoton'],
    "Biotech/Biology":  [r'\bbiotech', r'\bbiology\b', r'\bbiological\b', r'\bgenetic', r'\bgenome', r'\bprotein\b', r'\bcell\b', r'\bbacterial'],
    "Semiconductors":   [r'\bsemiconductor', r'\bsilicon\b', r'\bgallium', r'\bdoping\b', r'\bwafer'],
    "Robotics":         [r'\brobot', r'\bautonomo', r'\bmanipulator', r'\bactuat'],
    "Computers/Computing": [r'\bcomput', r'\bprocessor', r'\bsoftware\b', r'\bhardware\b', r'\bwireless\b', r'\bdigital\b', r'\bmicroprocessor'],
    "Material Science": [r'\bmaterial science', r'\bcomposite\b', r'\balloy\b', r'\bpolymer\b', r'\bceramics?\b', r'\bcoating'],
    "Cleantech":        [r'\bcleantech\b', r'\brenewable', r'\bsolar\b', r'\bwind energy', r'\bclean energy', r'\bgreen tech', r'\bhydrogen fuel'],
    "Hypertext":        [r'\bhypertext\b', r'\bhyperlink', r'\bweb page'],
    "Internet":         [r'\binternet\b', r'\bonline\b', r'\bworld wide web\b', r'\bbroadband', r'\bnetwork\b',
                         r'\bdot\.com\b', r'\bdotcom\b', r'\bdot com\b', r'\bdot-com\b'],
    "Chemistry":        [r'\bchemi', r'\breaction\b', r'\bcatalys', r'\bcompound\b', r'\bsynthes'],
    "Physics":          [r'\bphysics\b', r'\bphysical\b', r'\bquantum\b', r'\bthermodynamic', r'\bmechanics\b', r'\belectromagnet'],
    "Engineering":      [r'\bengineering\b'],
    "Nanotech":         [r'\bnanotech\w*'],
    "Nano":             [r'\bnano\w*'],
    "Molecular Manufacturing": [
        r'\bATP synthase\b', r'\bmolecular motor', r'\bmolecular machine',
        r'\bmolecular assembl', r'\bmolecular manufactur', r'\bconveyor',
        r'\bself-assembl', r'\bmolecular machinery\b', r'\bbiological machine',
        r'\bbiological motor', r'\bmolecular chaperones?\b', r'\bconveyor belt',
        r'\brobot arm', r'\bbiological nanomachin', r"nature's assembler",
        r'\brobotic assembly\b', r'\bmolecular robot',
    ],
    "Revolution":       [r'\brevolution', r'\brevolutionary\b', r'\bparadigm shift', r'\bbreakthrough', r'\btransformative\b'],
    "Financial":        [r'\$', r'\bdollar', r'\bmoney\b', r'\bfunding\b', r'\binvest', r'€', r'£', r'¥'],
    "Commerce":         [r'\bcommerce', r'\bcommercial', r'\bmarket', r'\btrade'],
    "Application":      [r'\bproduct', r'\bproduced', r'\bdevice', r'\bapplication', r'\bmanufactur'],
}


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def extract_date(text):
    """Find the first valid date in text."""
    for pattern, fmt in DATE_PATTERNS:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            candidate = match.group(1)
            parts = re.split(r'[\s,]+', candidate)
            has_month = any(p.lower() in MONTH_NAMES for p in parts)
            if not has_month and fmt != '%Y-%m-%d':
                continue
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                pass
            # Handle odd capitalization like "JUne" → "June"
            try:
                return datetime.strptime(candidate.title(), fmt)
            except ValueError:
                continue
    return None


def count_keyword(text, patterns):
    return sum(len(re.findall(p, text, re.IGNORECASE)) for p in patterns)


def count_words(text):
    return len(text.split())


def get_non_blank_lines(text):
    return [l.strip() for l in text.splitlines() if l.strip()]


def strip_references(text):
    """Split text into (body, references) at the first references heading."""
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if REFERENCES_START.match(line.strip()):
            return "\n".join(lines[:i]).strip(), "\n".join(lines[i:]).strip()
    return text, ""


def in_year_range(year):
    if year is None:
        return True
    return YEAR_MIN <= year <= YEAR_MAX


def _compute_date_match(original, scraped):
    """Compare two dates and return a match percentage.
    Year match is weighted heavily (90% minimum) since correct year
    is the most important factor for temporal analysis."""
    if original is None or scraped is None:
        return None
    if original.date() == scraped.date():
        return 100
    if original.year == scraped.year and original.month == scraped.month:
        return 95
    if original.year == scraped.year:
        return 90
    return 0


def _compute_days_between(original, scraped):
    """Absolute day gap between the two dates, or None if either is missing."""
    if original is None or scraped is None:
        return None
    return abs((original.date() - scraped.date()).days)


def _compute_date_match_tolerant(original, scraped):
    """Linear decay to 90 days: tolerates publication-vs-written gaps
    that straddle a year boundary (e.g. Dec 15 1994 vs Jan 5 1995 = 67%).
    Returns None if either date is missing."""
    days = _compute_days_between(original, scraped)
    if days is None:
        return None
    return max(0, round(100 * (1 - days / 90)))


def make_row(original_date, scraped_date, source_num, source_name, title, body, refs,
             original_year=None):
    """Build a single output row dict."""
    # Use reference year when available, otherwise derive from dates
    if original_year is not None:
        year = original_year
    elif original_date:
        year = original_date.year
    elif scraped_date:
        year = scraped_date.year
    else:
        year = None
    row = {
        'Original Date': original_date.strftime('%d %B %Y') if original_date else None,
        'Scraped Date':  scraped_date.strftime('%d %B %Y') if scraped_date else None,
        'Date Match %':  _compute_date_match(original_date, scraped_date),
        'Days Between':  _compute_days_between(original_date, scraped_date),
        'Date Match % (tolerant)': _compute_date_match_tolerant(original_date, scraped_date),
        'Year':          year,
        'Sources':       source_num,
        'Name':          source_name,
        'Word count':    count_words(body),
        'Title':         title,
        'Body':          body,
        'References':    refs,
    }
    for topic, patterns in TOPIC_KEYWORDS.items():
        row[topic] = count_keyword(body, patterns)
    # Combined column for electronics/computing/semiconductors
    row['Total Electronics/Computing'] = (
        row.get('Computers/Computing', 0) + row.get('Semiconductors', 0) + row.get('Electronics', 0)
    )
    return row


def _body_after_title(chunk, title):
    """Return the text in chunk after the first line matching title."""
    raw_lines = chunk.splitlines()
    for i, line in enumerate(raw_lines):
        if line.strip() == title:
            return "\n".join(raw_lines[i + 1:]).strip()
    return chunk


# ═════════════════════════════════════════════════════════════════════════════
# FORMAT-SPECIFIC PARSERS
# ═════════════════════════════════════════════════════════════════════════════

def _parse_ref_date(val):
    """Parse a single date value from the reference worksheet."""
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        fixed = val.replace('0ct', 'Oct').replace('0CT', 'OCT')
        try:
            return pd.to_datetime(fixed).to_pydatetime()
        except Exception:
            print(f"  WARNING: Could not parse date string: '{val}'")
            return None
    return None


def load_reference_data(script_dir, source_num):
    """Load dates and years from the reference worksheet for a given source.
    Returns (dates_list, years_list) where each is positionally indexed."""
    ref_path = script_dir / REFERENCE_WORKSHEET
    if not ref_path.exists():
        return [], []
    try:
        df = pd.read_excel(ref_path)
        src_df = df[df['Sources'] == source_num].reset_index(drop=True)
        dates = [_parse_ref_date(val) for val in src_df['Date']]
        years = []
        for val in src_df['Year']:
            if pd.isna(val):
                years.append(None)
            else:
                years.append(int(val))
        return dates, years
    except Exception as e:
        print(f"  WARNING: Could not load reference data: {e}")
        return [], []


def load_reference_dates(script_dir, source_num):
    """Load dates from the reference worksheet (backward-compatible wrapper)."""
    dates, _ = load_reference_data(script_dir, source_num)
    return dates


def count_reference_by_source(script_dir):
    """Return {source_num: row_count} from the reference worksheet.
    Used by the Coverage Report sheet to compare against output.xlsx."""
    ref_path = script_dir / REFERENCE_WORKSHEET
    if not ref_path.exists():
        return {}
    try:
        df = pd.read_excel(ref_path)
        return {int(k): int(v) for k, v in df.groupby('Sources').size().items()}
    except Exception as e:
        print(f"  WARNING: Could not count reference data: {e}")
        return {}


def align_by_date(scraped_dates, ref_dates, ref_years):
    """Align parsed articles to reference rows using date-group matching.

    Groups both sequences by date, then matches within each group by order.
    Returns (aligned_dates, aligned_years) lists parallel to scraped_dates.
    """
    from collections import defaultdict

    # Build reference groups: date -> [(ref_idx, ref_date, ref_year), ...]
    ref_groups = defaultdict(list)
    for i, (rd, ry) in enumerate(zip(ref_dates, ref_years)):
        if rd is not None:
            ref_groups[rd.date()].append((i, rd, ry))

    aligned_dates = [None] * len(scraped_dates)
    aligned_years = [None] * len(scraped_dates)

    # Track how many articles we've consumed from each date group
    group_cursors = defaultdict(int)

    for out_idx, sd in enumerate(scraped_dates):
        if sd is None:
            continue
        key = sd.date()
        cursor = group_cursors[key]
        group = ref_groups.get(key, [])
        if cursor < len(group):
            _, ref_date, ref_year = group[cursor]
            aligned_dates[out_idx] = ref_date
            aligned_years[out_idx] = ref_year
            group_cursors[key] = cursor + 1
        else:
            # No more ref rows for this date; use scraped date's year
            aligned_years[out_idx] = sd.year

    return aligned_dates, aligned_years


def parse_government(content, source_name, source_num, ref_dates=None, ref_years=None, stats=None):
    """Government: asterisk separators, strip 'Article N' prefix from title.
    Dates come from the reference worksheet (articles have no inline dates)."""
    chunks = [c.strip() for c in SEP_ASTERISKS.split(content) if c.strip()]
    if stats is not None:
        stats['raw_chunks'] = stats.get('raw_chunks', 0) + len(chunks)
    rows = []

    for idx, chunk in enumerate(chunks):
        original_date = ref_dates[idx] if ref_dates and idx < len(ref_dates) else None
        original_year = ref_years[idx] if ref_years and idx < len(ref_years) else None
        scraped_date = extract_date(chunk)
        date = original_date or scraped_date
        year = original_year or (date.year if date else None)

        # Skip year filter when using reference dates (reference is authoritative)
        if not ref_dates and not in_year_range(year):
            if stats is not None:
                stats['dropped_year'] = stats.get('dropped_year', 0) + 1
            continue

        lines = get_non_blank_lines(chunk)
        if not lines:
            if stats is not None:
                stats['dropped_empty_chunk'] = stats.get('dropped_empty_chunk', 0) + 1
            continue

        # Skip "Article N" prefix line
        start = 0
        if re.match(r'^Article\s+\d+\s*$', lines[0], re.IGNORECASE):
            start = 1

        if start >= len(lines):
            if stats is not None:
                stats['dropped_empty_title'] = stats.get('dropped_empty_title', 0) + 1
            continue

        title = lines[start]
        body_text = _body_after_title(chunk, title)
        body, refs = strip_references(body_text)

        # Fix: if body is empty but title is very long, the body was pasted into the title
        if count_words(body) == 0 and count_words(title) > 20:
            body = title
            title = title[:80].rsplit(' ', 1)[0] + '...'

        rows.append(make_row(original_date, scraped_date, source_num, source_name, title, body, refs,
                             original_year=original_year))

    if stats is not None:
        stats['output_rows'] = stats.get('output_rows', 0) + len(rows)
    return rows


def parse_after_label(content, source_name, source_num, ref_dates=None, ref_years=None, stats=None):
    """Science Research / Science News: section label precedes title."""
    chunks = [c.strip() for c in SEP_ASTERISKS.split(content) if c.strip()]
    if stats is not None:
        stats['raw_chunks'] = stats.get('raw_chunks', 0) + len(chunks)
    rows, skipped = [], 0

    for idx, chunk in enumerate(chunks):
        scraped_date = extract_date(chunk)
        original_date = ref_dates[idx] if ref_dates and idx < len(ref_dates) else None
        original_year = ref_years[idx] if ref_years and idx < len(ref_years) else None
        date = scraped_date or original_date
        year = date.year if date else None
        if not in_year_range(year):
            skipped += 1
            if stats is not None:
                stats['dropped_year'] = stats.get('dropped_year', 0) + 1
            continue

        lines = get_non_blank_lines(chunk)
        if not lines:
            if stats is not None:
                stats['dropped_empty_chunk'] = stats.get('dropped_empty_chunk', 0) + 1
            continue

        # Find title: line after a section label
        title = ""
        for i, line in enumerate(lines):
            if SECTION_LABELS.match(line) and i + 1 < len(lines):
                candidate = lines[i + 1]
                if not SCIENCE_META.match(candidate) and not re.match(r'^\d+$', candidate):
                    title = candidate
                    break

        # Fallback: first substantive non-metadata line
        if not title:
            for line in lines:
                if (not SCIENCE_META.match(line)
                        and not re.match(r'^\d+$', line)
                        and len(line) > 3):
                    title = line
                    break

        if not title:
            if stats is not None:
                stats['dropped_empty_title'] = stats.get('dropped_empty_title', 0) + 1
            continue

        # Body after title, strip leading science metadata
        body_text = _body_after_title(chunk, title)
        body_lines = body_text.splitlines()
        while body_lines and (not body_lines[0].strip()
                              or SCIENCE_META.match(body_lines[0].strip())):
            body_lines.pop(0)

        body_text = "\n".join(body_lines).strip()
        body, refs = strip_references(body_text)

        # Fix: if body is empty but title is very long, the content ended up as the title
        if count_words(body) < 5 and count_words(title) > 20:
            body = title
            title = title[:80].rsplit(' ', 1)[0] + '...'

        if count_words(body) < 5:
            if stats is not None:
                stats['dropped_wordcount'] = stats.get('dropped_wordcount', 0) + 1
            continue

        rows.append(make_row(original_date, scraped_date, source_num, source_name, title, body, refs,
                             original_year=original_year))

    if skipped:
        print(f"  (skipped {skipped} articles outside {YEAR_MIN}–{YEAR_MAX})")
    if stats is not None:
        stats['output_rows'] = stats.get('output_rows', 0) + len(rows)
    return rows


def _strip_business_tail(lines):
    """Remove trailing junk (Document ID, 'More Like This', contact blocks)."""
    while lines:
        last = lines[-1].strip()
        if not last or BUSINESS_TAIL_JUNK.match(last):
            lines.pop()
        else:
            break
    # Strip trailing contact block (has both phone AND email on same line)
    if lines and (re.search(r'\d{3}[.-]\d{3}[.-]\d{4}', lines[-1])
                  and re.search(r'\b\S+@\S+\.\S+', lines[-1])):
        lines.pop()
    return lines


def parse_business(content, source_name, source_num, ref_dates=None, ref_years=None, stats=None):
    """Business Press / Business: 'Article N ****' separator, real headline as title."""
    chunks = [c.strip() for c in SEP_BUSINESS.split(content) if c.strip()]
    if stats is not None:
        stats['raw_chunks'] = stats.get('raw_chunks', 0) + len(chunks)
    rows, skipped = [], 0

    for idx, chunk in enumerate(chunks):
        scraped_date = extract_date(chunk)
        original_date = ref_dates[idx] if ref_dates and idx < len(ref_dates) else None
        original_year = ref_years[idx] if ref_years and idx < len(ref_years) else None
        date = scraped_date or original_date
        year = date.year if date else None
        if not in_year_range(year):
            skipped += 1
            if stats is not None:
                stats['dropped_year'] = stats.get('dropped_year', 0) + 1
            continue

        lines = get_non_blank_lines(chunk)
        if not lines:
            if stats is not None:
                stats['dropped_empty_chunk'] = stats.get('dropped_empty_chunk', 0) + 1
            continue

        # Skip known header junk at start of chunk
        while lines and BUSINESS_HEADER_JUNK.match(lines[0]):
            lines.pop(0)
        if not lines:
            if stats is not None:
                stats['dropped_empty_title'] = stats.get('dropped_empty_title', 0) + 1
            continue

        title = lines[0]  # actual headline

        # Body starts after copyright line
        raw_lines = chunk.splitlines()
        body_start = 0
        for i, line in enumerate(raw_lines):
            if COPYRIGHT_LINE.match(line.strip()):
                body_start = i + 1
                break

        if body_start == 0:
            # Fallback: skip metadata block (~first 8 lines)
            body_start = min(8, len(raw_lines))

        body_lines = list(raw_lines[body_start:])
        body_lines = _strip_business_tail(body_lines)
        body_text = "\n".join(body_lines).strip()
        body, refs = strip_references(body_text)

        if count_words(body) < 5:
            if stats is not None:
                stats['dropped_wordcount'] = stats.get('dropped_wordcount', 0) + 1
            continue

        rows.append(make_row(original_date, scraped_date, source_num, source_name, title, body, refs,
                             original_year=original_year))

    if skipped:
        print(f"  (skipped {skipped} articles outside {YEAR_MIN}–{YEAR_MAX})")
    if stats is not None:
        stats['output_rows'] = stats.get('output_rows', 0) + len(rows)
    return rows


def parse_futurist(content, source_name, source_num, ref_dates=None, ref_years=None, stats=None):
    """Futurists: split by ToC lines, asterisk separators, and dash separators."""
    combined_sep = re.compile(
        r'(?:^\s*Foresight Update \d+\s*-\s*Table of Contents.*$'
        r'|^\s*\*{19,}\s*$'
        r'|^\s*-{40,}\s*$)',
        re.MULTILINE
    )
    chunks = [c.strip() for c in combined_sep.split(content) if c.strip()]
    if stats is not None:
        stats['raw_chunks'] = stats.get('raw_chunks', 0) + len(chunks)

    # Lines that appear in issue headers (not article content)
    header_line_re = re.compile(
        r'^(A publication of the Foresight Institute|'
        r'Preparing for future technologies|'
        r'Board of Directors|'
        r'All Rights Reserved|'
        r'Write to the Foresight Institute|'
        r'If you find information|'
        r'Editor\s|Publisher\s|'
        r'.{0,30}(President|Secretary|Treasurer)\s*$|'
        r'Box \d+.*CA\s+\d|'
        r'.{0,3}Copyright\s+\d{4}|'
        r'Foresight Institute\s*$|'
        r'\d{1,2}\s+\w+\s+\d{4}\s*$)',
        re.IGNORECASE
    )

    # Whole chunks to skip entirely
    skip_chunk_re = re.compile(
        r'^(Clippings Invited|'
        r'If you find information and clippings|'
        r'Write to the Foresight Institute)',
        re.IGNORECASE
    )

    # Detect issue header chunks to reset date propagation
    issue_header_re = re.compile(
        r'^\s*A publication of the Foresight Institute',
        re.IGNORECASE
    )

    rows, skipped = [], 0
    last_date = None
    article_idx = 0

    for chunk in chunks:
        lines = get_non_blank_lines(chunk)
        if not lines:
            continue

        # Reset last_date at issue boundaries to prevent stale date propagation
        if issue_header_re.match(lines[0]):
            last_date = None

        # Always try to capture date (even from skipped header chunks)
        chunk_date = extract_date(chunk)
        if chunk_date:
            last_date = chunk_date

        # Skip known boilerplate chunks
        if any(skip_chunk_re.match(l) for l in lines[:2]):
            if stats is not None:
                stats['dropped_boilerplate'] = stats.get('dropped_boilerplate', 0) + 1
            continue
        if count_words(chunk) < 20:
            if stats is not None:
                stats['dropped_boilerplate'] = stats.get('dropped_boilerplate', 0) + 1
            continue

        # Strip issue-header boilerplate from chunk start (only when
        # the chunk actually begins with a known header line)
        if lines and header_line_re.match(lines[0]):
            while lines:
                if header_line_re.match(lines[0]):
                    lines.pop(0)
                elif len(lines[0].split()) <= 2:
                    lines.pop(0)  # short name lines in header block
                elif lines[0][0].islower():
                    lines.pop(0)  # continuation of previous header paragraph
                else:
                    break

        if not lines or count_words(' '.join(lines)) < 15:
            if stats is not None:
                stats['dropped_boilerplate'] = stats.get('dropped_boilerplate', 0) + 1
            continue

        scraped_date = chunk_date if chunk_date else last_date
        year = scraped_date.year if scraped_date else None
        if not in_year_range(year):
            skipped += 1
            if stats is not None:
                stats['dropped_year'] = stats.get('dropped_year', 0) + 1
            continue

        original_date = ref_dates[article_idx] if ref_dates and article_idx < len(ref_dates) else None
        original_year = ref_years[article_idx] if ref_years and article_idx < len(ref_years) else None
        title = lines[0]
        body = _body_after_title(chunk, title)
        rows.append(make_row(original_date, scraped_date, source_num, source_name, title, body, "",
                             original_year=original_year))
        article_idx += 1

    if skipped:
        print(f"  (skipped {skipped} articles outside {YEAR_MIN}–{YEAR_MAX})")
    if stats is not None:
        stats['output_rows'] = stats.get('output_rows', 0) + len(rows)
    return rows


def parse_newspaper(content, source_name, source_num, ref_dates=None, ref_years=None, stats=None):
    """Newspapers: split by 'Article N ****' and '* * *' sub-item separators.
    Uses date-group alignment instead of positional index for reference matching."""
    chunks = [c.strip() for c in SEP_NEWSPAPER.split(content) if c.strip()]
    if stats is not None:
        stats['raw_chunks'] = stats.get('raw_chunks', 0) + len(chunks)
    skipped = 0
    last_date = None

    # First pass: parse all chunks into preliminary article data
    parsed = []
    for idx, chunk in enumerate(chunks):
        scraped_date = extract_date(chunk)
        if scraped_date:
            last_date = scraped_date
        else:
            scraped_date = last_date

        date = scraped_date
        year = date.year if date else None
        if not in_year_range(year):
            skipped += 1
            if stats is not None:
                stats['dropped_year'] = stats.get('dropped_year', 0) + 1
            continue

        lines = get_non_blank_lines(chunk)
        if not lines:
            if stats is not None:
                stats['dropped_empty_chunk'] = stats.get('dropped_empty_chunk', 0) + 1
            continue

        # Skip known header junk
        while lines and BUSINESS_HEADER_JUNK.match(lines[0]):
            lines.pop(0)
        if not lines:
            if stats is not None:
                stats['dropped_empty_title'] = stats.get('dropped_empty_title', 0) + 1
            continue

        title = lines[0]

        # Body starts after copyright line if present
        raw_lines = chunk.splitlines()
        body_start = 0
        for i, line in enumerate(raw_lines):
            if COPYRIGHT_LINE.match(line.strip()):
                body_start = i + 1
                break

        if body_start == 0:
            body_text = _body_after_title(chunk, title)
        else:
            body_lines = list(raw_lines[body_start:])
            body_lines = _strip_business_tail(body_lines)
            body_text = "\n".join(body_lines).strip()

        body, refs = strip_references(body_text)

        if count_words(body) < 5:
            if stats is not None:
                stats['dropped_wordcount'] = stats.get('dropped_wordcount', 0) + 1
            continue

        parsed.append({
            'scraped_date': scraped_date,
            'title': title,
            'body': body,
            'refs': refs,
        })

    # Second pass: align with reference data by date groups
    scraped_dates = [p['scraped_date'] for p in parsed]
    if ref_dates and ref_years:
        aligned_dates, aligned_years = align_by_date(scraped_dates, ref_dates, ref_years)
    else:
        aligned_dates = [None] * len(parsed)
        aligned_years = [None] * len(parsed)

    # Build final rows with aligned reference data
    rows = []
    for i, p in enumerate(parsed):
        rows.append(make_row(
            aligned_dates[i], p['scraped_date'], source_num, source_name,
            p['title'], p['body'], p['refs'],
            original_year=aligned_years[i],
        ))

    if skipped:
        print(f"  (skipped {skipped} articles outside {YEAR_MIN}–{YEAR_MAX})")
    if stats is not None:
        stats['output_rows'] = stats.get('output_rows', 0) + len(rows)
    return rows


# ═════════════════════════════════════════════════════════════════════════════
# DISPATCHER
# ═════════════════════════════════════════════════════════════════════════════

def strip_rtf(text):
    """Strip RTF markup, returning plain text."""
    # Remove RTF header up to first empty line or content
    text = re.sub(r'\{\\fonttbl[^}]*\}', '', text)
    text = re.sub(r'\{\\colortbl[^}]*\}', '', text)
    text = re.sub(r'\{\\rtf1[^}]*?\n', '', text)
    # Remove RTF control words (e.g., \f0, \fs24, \cf0, \pard, etc.)
    text = re.sub(r'\\[a-z]+\d*\s?', ' ', text)
    # Remove remaining braces
    text = text.replace('{', '').replace('}', '')
    # Convert RTF line breaks (\) to newlines
    text = text.replace('\\\n', '\n')
    # Clean up extra whitespace
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n[ \t]+', '\n', text)
    return text.strip()


def parse_articles(filepath, source_name, source_num, title_style, stats=None):
    """Route to the appropriate parser. If stats is passed, parser counters
    (raw_chunks, dropped_*, output_rows) are accumulated in place."""
    script_dir = filepath.parent if filepath.suffix != '.rtf' else filepath.parent.parent

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()
    content = content.replace('\r\n', '\n').replace('\r', '\n')

    # Strip RTF markup if needed
    if filepath.suffix == '.rtf':
        content = strip_rtf(content)

    # Skip reference dates for supplementary files (e.g., Business_2005.rtf)
    # whose articles don't align with the main reference worksheet
    if filepath.suffix == '.rtf':
        ref_dates, ref_years = [], []
    else:
        ref_dates, ref_years = load_reference_data(script_dir, source_num)
    if ref_dates:
        print(f"  (loaded {len(ref_dates)} reference entries from {REFERENCE_WORKSHEET})")

    if title_style == "government":
        return parse_government(content, source_name, source_num, ref_dates, ref_years, stats=stats)
    elif title_style == "after_label":
        return parse_after_label(content, source_name, source_num, ref_dates, ref_years, stats=stats)
    elif title_style == "business_press":
        return parse_business(content, source_name, source_num, ref_dates, ref_years, stats=stats)
    elif title_style == "futurist":
        return parse_futurist(content, source_name, source_num, ref_dates, ref_years, stats=stats)
    elif title_style == "newspaper":
        return parse_newspaper(content, source_name, source_num, ref_dates, ref_years, stats=stats)
    else:
        raise ValueError(f"Unknown title_style: {title_style}")


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═════════════════════════════════════════════════════════════════════════════

def write_excel(all_rows, output_path):
    columns = (
        ['Original Date', 'Scraped Date', 'Date Match %',
         'Days Between', 'Date Match % (tolerant)',
         'Year', 'Sources', 'Name',
         'Word count', 'Title', 'Body', 'References']
        + list(TOPIC_KEYWORDS.keys())
        + ['Total Electronics/Computing']
    )
    df = pd.DataFrame(all_rows, columns=columns)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        ws = writer.sheets['Sheet1']

        header_fill = PatternFill('solid', start_color='4472C4', end_color='4472C4')
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        name_widths = {
            'Original Date': 14, 'Scraped Date': 14, 'Date Match %': 12,
            'Days Between': 10, 'Date Match % (tolerant)': 18,
            'Year': 6, 'Sources': 8, 'Name': 18, 'Word count': 12,
            'Title': 40, 'Body': 60, 'References': 60,
        }
        all_cols = [chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26)
                    for i in range(ws.max_column)]
        for col_idx, col_letter in enumerate(all_cols, start=1):
            col_name = df.columns[col_idx - 1]
            ws.column_dimensions[col_letter].width = name_widths.get(col_name, 16)

        data_font = Font(name='Arial', size=11)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = Alignment(wrap_text=False, vertical='top')


def write_qualitative_sheets(output_path):
    """Append 'Qualitative Samples' + 'AI Qualitative Samples' sheets.

    Uses the sampling/ package so the pipeline and the notebook share one
    implementation (the previous flow required the user to run the bottom
    notebook cells manually — this runs automatically).
    """
    from sampling.qualitative import build_qualitative_samples
    from sampling.ai_qualitative import build_ai_qualitative_samples

    df_all = pd.read_excel(output_path, sheet_name='Sheet1')
    qual_df = build_qualitative_samples(df_all, year_min=YEAR_MIN, year_max=YEAR_MAX)
    ai_df = build_ai_qualitative_samples(df_all, source_filter=6,
                                          year_min=YEAR_MIN, year_max=YEAR_MAX)

    with pd.ExcelWriter(output_path, engine='openpyxl',
                        mode='a', if_sheet_exists='replace') as writer:
        qual_df.to_excel(writer, sheet_name='Qualitative Samples', index=False)
        ai_df.to_excel(writer, sheet_name='AI Qualitative Samples', index=False)

    print(f"Wrote {len(qual_df)} qualitative + {len(ai_df)} AI-qualitative samples.")


# Source-num → human label, matches FILE_SOURCE_MAP. Kept as a module-level
# lookup so the Coverage Report can label rows nicely.
SOURCE_LABELS = {
    1: 'Government',
    2: 'Science News',
    3: 'Science Research',
    4: 'Business Press',
    5: 'Business',
    6: 'Futurists',
    7: 'Newspapers',
}

# Sources where reference-count alignment is approximate (Futurists uses an
# ad-hoc sequential counter, so its per-source stats can't be compared 1-to-1
# against the reference worksheet). Flagged in the Coverage Report.
APPROXIMATE_SOURCES = {6}


def write_coverage_sheet(coverage_stats, ref_counts, output_path):
    """Append a 'Coverage Report' sheet explaining the per-source article-count
    delta between the reference worksheet and output.xlsx.

    `coverage_stats` is a dict keyed by source_num with the counters each
    parser accumulates (raw_chunks, dropped_year, dropped_wordcount, etc.)
    plus post-hoc fields (dropped_dedup, final_count) added by main().
    """
    rows = []
    for src_num in sorted(set(list(coverage_stats.keys()) + list(ref_counts.keys()))):
        s = coverage_stats.get(src_num, {})
        label = SOURCE_LABELS.get(src_num, f'Source {src_num}')
        if src_num in APPROXIMATE_SOURCES:
            label += ' (approx)'
        ref = ref_counts.get(src_num, 0)
        final = s.get('final_count', 0)
        rows.append({
            'Source': label,
            'Sources#': src_num,
            'Reference Count': ref,
            'Raw Chunks': s.get('raw_chunks', 0),
            'Dropped (Year)': s.get('dropped_year', 0),
            'Dropped (Word Count)': s.get('dropped_wordcount', 0),
            'Dropped (Empty Title)': s.get('dropped_empty_title', 0),
            'Dropped (Empty Chunk)': s.get('dropped_empty_chunk', 0),
            'Dropped (Boilerplate)': s.get('dropped_boilerplate', 0),
            'Dropped (Dedup)': s.get('dropped_dedup', 0),
            'Final Count': final,
            'Delta': final - ref,
            'Coverage %': round(100 * final / ref, 1) if ref else None,
        })

    # Totals row
    total_ref = sum(r['Reference Count'] for r in rows)
    total_final = sum(r['Final Count'] for r in rows)
    rows.append({
        'Source': 'TOTAL',
        'Sources#': '',
        'Reference Count': total_ref,
        'Raw Chunks': sum(r['Raw Chunks'] for r in rows),
        'Dropped (Year)': sum(r['Dropped (Year)'] for r in rows),
        'Dropped (Word Count)': sum(r['Dropped (Word Count)'] for r in rows),
        'Dropped (Empty Title)': sum(r['Dropped (Empty Title)'] for r in rows),
        'Dropped (Empty Chunk)': sum(r['Dropped (Empty Chunk)'] for r in rows),
        'Dropped (Boilerplate)': sum(r['Dropped (Boilerplate)'] for r in rows),
        'Dropped (Dedup)': sum(r['Dropped (Dedup)'] for r in rows),
        'Final Count': total_final,
        'Delta': total_final - total_ref,
        'Coverage %': round(100 * total_final / total_ref, 1) if total_ref else None,
    })

    df = pd.DataFrame(rows)
    with pd.ExcelWriter(output_path, engine='openpyxl',
                        mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Coverage Report', index=False)
        ws = writer.sheets['Coverage Report']
        header_fill = PatternFill('solid', start_color='4472C4', end_color='4472C4')
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        # Bold the TOTAL row (last row)
        total_row_idx = ws.max_row
        total_font = Font(name='Arial', bold=True, size=11)
        for cell in ws[total_row_idx]:
            cell.font = total_font
        # Widen columns
        widths = {'A': 22, 'B': 10, 'C': 16, 'D': 12}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        for col_letter in 'EFGHIJKLMN':
            ws.column_dimensions[col_letter].width = 16

    print(f"Wrote Coverage Report ({total_final}/{total_ref} articles, "
          f"{100*total_final/total_ref:.1f}% coverage)" if total_ref else "")


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

def main():
    from collections import defaultdict

    script_dir = Path(__file__).parent
    all_rows = []
    # Per-source stats accumulated by the parsers; used for the Coverage Report.
    coverage_stats = defaultdict(lambda: {
        'raw_chunks': 0,
        'dropped_year': 0,
        'dropped_wordcount': 0,
        'dropped_empty_title': 0,
        'dropped_empty_chunk': 0,
        'dropped_boilerplate': 0,
        'output_rows': 0,
    })

    for filename, (source_name, source_num, title_style) in FILE_SOURCE_MAP.items():
        filepath = script_dir / filename
        if not filepath.exists():
            print(f"WARNING: '{filename}' not found in {script_dir} — skipping.")
            continue

        rows = parse_articles(filepath, source_name, source_num, title_style,
                              stats=coverage_stats[source_num])
        all_rows.extend(rows)
        print(f"'{filename}' ({source_name}, Sources={source_num}): {len(rows)} articles")
        # Print a few samples for spot-checking
        for r in rows[:3]:
            safe_title = r['Title'][:60].encode('ascii', 'replace').decode('ascii')
            print(f"  Sample: Scraped={r['Scraped Date']}, Original={r['Original Date']}, "
                  f"Match={r['Date Match %']}, Words={r['Word count']}, "
                  f"Title='{safe_title}'")

    if not all_rows:
        print("No articles found. Check that your source files are in the same folder.")
        return

    # ── DEDUPLICATION ───────────────────────────────────────────────────────
    before_count = len(all_rows)
    df_temp = pd.DataFrame(all_rows)
    # Compute per-source dup counts *before* dropping so the Coverage Report
    # can attribute dedup loss to the right source.
    dup_mask = df_temp.duplicated(subset=['Title', 'Body', 'Sources'], keep='first')
    for src_num, n in df_temp.loc[dup_mask].groupby('Sources').size().items():
        coverage_stats[int(src_num)]['dropped_dedup'] = int(n)
    df_temp = df_temp.drop_duplicates(subset=['Title', 'Body', 'Sources'], keep='first')
    all_rows = df_temp.to_dict('records')
    removed = before_count - len(all_rows)
    if removed:
        print(f"\nRemoved {removed} exact duplicate rows ({before_count} -> {len(all_rows)})")

    # Record final post-dedup counts per source for the Coverage Report.
    for src_num, n in df_temp.groupby('Sources').size().items():
        coverage_stats[int(src_num)]['final_count'] = int(n)

    # ── DATE MATCH REPORTING (exclude Government & Futurists) ───────────────
    def _valid(v):
        if v is None:
            return False
        if isinstance(v, float) and v != v:  # NaN
            return False
        return True

    print("\n-- Date Match by Source (strict vs tolerant) --")
    for src_num, src_name in [(2, 'Science News'), (3, 'Science Research'),
                               (4, 'Business Press'), (5, 'Business'), (7, 'Newspapers')]:
        strict = [r['Date Match %'] for r in all_rows
                  if r['Sources'] == src_num and _valid(r['Date Match %'])]
        tolerant = [r['Date Match % (tolerant)'] for r in all_rows
                    if r['Sources'] == src_num and _valid(r['Date Match % (tolerant)'])]
        days = [r['Days Between'] for r in all_rows
                if r['Sources'] == src_num and _valid(r['Days Between'])]
        if strict:
            s_avg = sum(strict) / len(strict)
            t_avg = sum(tolerant) / len(tolerant) if tolerant else float('nan')
            d_median = sorted(days)[len(days) // 2] if days else None
            s_zeros = sum(1 for v in strict if v == 0)
            t_zeros = sum(1 for v in tolerant if v == 0)
            print(f"  {src_name}: strict={s_avg:.1f}%  tolerant={t_avg:.1f}%  "
                  f"median_days={d_median}  n={len(strict)}")
            if s_zeros != t_zeros:
                print(f"    (tolerant reduces 0%-rows from {s_zeros} -> {t_zeros})")

    output_path = script_dir / OUTPUT_FILE
    write_excel(all_rows, output_path)
    print(f"\nDone! {len(all_rows)} total rows written to '{output_path}'")

    # ── AUTO-POPULATE qualitative & AI-qualitative sample sheets ────────────
    try:
        write_qualitative_sheets(output_path)
    except Exception as e:
        print(f"  WARNING: could not write qualitative sample sheets: {e}")

    # ── COVERAGE REPORT ────────────────────────────────────────────────────
    try:
        ref_counts = count_reference_by_source(script_dir)
        write_coverage_sheet(dict(coverage_stats), ref_counts, output_path)
    except Exception as e:
        print(f"  WARNING: could not write Coverage Report: {e}")


if __name__ == '__main__':
    main()
