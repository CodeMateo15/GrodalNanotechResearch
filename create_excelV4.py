"""
Convert txt/rtf files into Excel format matching Combined_community_worksheet.xlsx.

V4: Adds Science News, Business RTF, Futurists support; fixes Business Press separator.

HOW TO USE:
1. Place this script in the same folder as your .txt/.rtf files
2. Run: python create_excelV4.py
"""

import re
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIGURATION ────────────────────────────────────────────────────────────

FILE_SOURCE_MAP = {
    "government.txt":       ("Government",        1, "government"),
    "Science_news.txt":     ("Science News",       2, "after_label"),
    "Science_research.txt": ("Science Research",   3, "after_label"),
    "Business_press.txt":   ("Business Press",     4, "business_press"),
    "Business_2005.rtf":    ("Business",           5, "business_rtf"),
    "futurists.txt":        ("Futurists",          6, "futurist"),
}

YEAR_MIN = 1983
YEAR_MAX = 2005

OUTPUT_FILE = "output.xlsx"

# Reference worksheet with dates for sources that lack them in the raw text
REFERENCE_WORKSHEET = "Combined community worksheet.xlsx"

# ── SEPARATORS ───────────────────────────────────────────────────────────────

SEP_ASTERISKS = re.compile(r'^\s*\*{5,}\s*$', re.MULTILINE)
SEP_BUSINESS  = re.compile(r'^\s*Article\s+\d+\s+\*{5,}', re.MULTILINE)

# ── SECTION LABELS (Science-style) ──────────────────────────────────────────

SECTION_LABELS = re.compile(
    r'^(Reports?|Policy\s+Forum|News|Perspectives?|Reviews?|Letters?|'
    r'Editorial|Research\s+Article|Brief\s+Communications?|Brevia|'
    r'Random\s+Samples?|Newsmakers?|News\s+Focus|'
    r'News\s+of\s+the\s+Week|Findings|'
    r'This\s+Week\s+in\s+Science|Essays?|Corrections?)$',
    re.IGNORECASE
)

# ── REFERENCES ───────────────────────────────────────────────────────────────

REFERENCES_START = re.compile(
    r'^(References(\s+and\s+Notes)?|Bibliography|Notes|'
    r'Supporting\s+(Online\s+)?Material|SOM\s+Text|'
    r'Acknowledgements?|Supplementary\s+Materials?)$',
    re.IGNORECASE
)

# ── BUSINESS PRESS ───────────────────────────────────────────────────────────

COPYRIGHT_LINE = re.compile(r'^\(?\s*Copyright|\(c\)\s*\d{4}', re.IGNORECASE)

BUSINESS_TAIL_JUNK = re.compile(
    r'^\s*(Document\s+\S+|More Like This)\s*$',
    re.IGNORECASE
)

# ── SCIENCE METADATA ────────────────────────────────────────────────────────

SCIENCE_META = re.compile(
    r'^(\d{1,5}\s*$|'
    r'DOI:\s|doi:\s|'
    r'Vol\.\s*\d|'
    r'Prev\s*\|\s*Table of Contents|'
    r'Science\s+\d{1,2}\s+\w+\s+\d{4}|'
    r'Science,\s+New\s+Series|'
    r'Copyright\s*\(?|'
    r'Originally\s+published|'
    r'\d{1,2}\s+\w+\s+\d{4}\s*$|'
    r'\d{1,2}\s+\w+\s+\d{4}\s+VOL\s+\d)',
    re.IGNORECASE
)

# ── DATES ────────────────────────────────────────────────────────────────────

DATE_PATTERNS = [
    (r'\b(\d{1,2}\s+\w+\s+\d{4})\b',    '%d %B %Y'),
    (r'\b(\w+\s+\d{1,2},\s+\d{4})\b',   '%B %d, %Y'),
    (r'\b(\d{1,2}\s+\w{3}\s+\d{4})\b',  '%d %b %Y'),
    (r'\b(\w{3}\s+\d{1,2},\s+\d{4})\b', '%b %d, %Y'),
    (r'\b(\d{4}-\d{2}-\d{2})\b',         '%Y-%m-%d'),
]

MONTH_NAMES = {
    'january','february','march','april','may','june',
    'july','august','september','october','november','december',
    'jan','feb','mar','apr','jun','jul','aug','sep','oct','nov','dec'
}

# ── TOPIC KEYWORDS ───────────────────────────────────────────────────────────

TOPIC_KEYWORDS = {
    "Space":            [r'\bspace\b', r'\bsatellite', r'\baerospace', r'\borbit', r'\brocket', r'\bnasa\b', r'\bspacecraft'],
    "Electronics":      [r'\belectronic', r'\bsemiconductor', r'\bcircuit', r'\btransistor', r'\bdiode', r'\bchip\b'],
    "Artificial Intelligence": [r'\bartificial intelligence\b', r'\bmachine learning\b', r'\bneural network', r'\bdeep learning\b', r'\bai\b', r'\bgradient descent\b'],
    "Photonics":        [r'\bphotonic', r'\boptical\b', r'\blaser', r'\bfiber optic', r'\bphoton'],
    "Biotech/Biology":  [r'\bbiotech', r'\bbiology\b', r'\bbiological\b', r'\bgenetic', r'\bgenome', r'\bprotein\b', r'\bcell\b', r'\bbacterial'],
    "Semiconductors":   [r'\bsemiconductor', r'\bsilicon\b', r'\bgallium', r'\bdoping\b', r'\bwafer'],
    "Robotics":         [r'\brobot', r'\bautonomo', r'\bmanipulator', r'\bactuat'],
    "Computers/Computing": [r'\bcomput', r'\bprocessor', r'\bsoftware\b', r'\bhardware\b', r'\bwireless\b', r'\bdigital\b', r'\bmicroprocessor'],
    "Material Science": [r'\bmaterial science', r'\bcomposite\b', r'\balloy\b', r'\bpolymer\b', r'\bceramics?\b', r'\bcoating'],
    "Cleantech":        [r'\bcleantech\b', r'\brenewable', r'\bsolar\b', r'\bwind energy', r'\bclean energy', r'\bgreen tech', r'\bhydrogen fuel'],
    "Hypertext":        [r'\bhypertext\b', r'\bhyperlink', r'\bhtml\b', r'\bweb page', r'\bwww\b'],
    "Internet":         [r'\binternet\b', r'\bonline\b', r'\bworld wide web\b', r'\bbroadband', r'\bnetwork\b'],
    "Chemistry":        [r'\bchemi', r'\bmolecul', r'\breaction\b', r'\bcatalys', r'\bcompound\b', r'\bsynthes'],
    "Physics":          [r'\bphysics\b', r'\bphysical\b', r'\bquantum\b', r'\bthermodynamic', r'\bmechanics\b', r'\belectromagnet'],
    "Nanotech":         [r'\bnanotech\w*'],
    "Nano":             [r'\bnano\w*'],
}


# ═════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def extract_date(text):
    """Find the first valid date in text."""
    for pattern, fmt in DATE_PATTERNS:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            candidate = match.group(1)
            parts = re.split(r'[\s,]+', candidate)
            has_month = any(p.lower() in MONTH_NAMES for p in parts)
            if not has_month and fmt != '%Y-%m-%d':
                continue
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                pass
            # Handle odd capitalization like "JUne" → "June"
            try:
                return datetime.strptime(candidate.title(), fmt)
            except ValueError:
                continue
    return None


def count_keyword(text, patterns):
    return sum(len(re.findall(p, text, re.IGNORECASE)) for p in patterns)


def count_words(text):
    return len(text.split())


def get_non_blank_lines(text):
    return [l.strip() for l in text.splitlines() if l.strip()]


def strip_references(text):
    """Split text into (body, references) at the first references heading."""
    lines = text.splitlines()
    for i, line in enumerate(lines):
        if REFERENCES_START.match(line.strip()):
            return "\n".join(lines[:i]).strip(), "\n".join(lines[i:]).strip()
    return text, ""


def in_year_range(year):
    if year is None:
        return True
    return YEAR_MIN <= year <= YEAR_MAX


def make_row(date, year, source_num, source_name, title, body, refs):
    """Build a single output row dict."""
    row = {
        'Date':       date.strftime('%d %B %Y') if date else None,
        'Year':       year,
        'Sources':    source_num,
        'Name':       source_name,
        'Word count': count_words(body),
        'Title':      title,
        'Body':       body,
        'References': refs,
    }
    for topic, patterns in TOPIC_KEYWORDS.items():
        row[topic] = count_keyword(body, patterns)
    return row


def _body_after_title(chunk, title):
    """Return the text in chunk after the first line matching title."""
    raw_lines = chunk.splitlines()
    for i, line in enumerate(raw_lines):
        if line.strip() == title:
            return "\n".join(raw_lines[i + 1:]).strip()
    return chunk


# ═════════════════════════════════════════════════════════════════════════════
# RTF STRIPPING
# ═════════════════════════════════════════════════════════════════════════════

def strip_rtf(text):
    """Convert RTF to plain text."""
    # Handle \'xx hex escapes (e.g. \'e9 → é)
    text = re.sub(r"\\'([0-9a-fA-F]{2})",
                  lambda m: chr(int(m.group(1), 16)), text)

    # Remove font/color table groups
    text = re.sub(r'\{\\fonttbl[^}]*\}', '', text)
    text = re.sub(r'\{\\colortbl[^}]*\}', '', text)

    # Remove {\rtf1... opening line
    text = re.sub(r'\{\\rtf1[^\n]*', '', text)

    # Preserve RTF-escaped characters
    text = text.replace('\\{', '\x00LB\x00')
    text = text.replace('\\}', '\x00RB\x00')
    text = text.replace('\\\\', '\x00BS\x00')

    # Replace \<newline> with newline (RTF line continuation)
    text = text.replace('\\\r\n', '\n')
    text = text.replace('\\\r', '\n')
    text = text.replace('\\\n', '\n')

    # Remove RTF control words (\word or \word123, optionally eating a trailing space)
    text = re.sub(r'\\[a-zA-Z]+\d*\s?', ' ', text)

    # Remove remaining braces (RTF grouping)
    text = text.replace('{', '').replace('}', '')

    # Restore escaped characters
    text = text.replace('\x00LB\x00', '{')
    text = text.replace('\x00RB\x00', '}')
    text = text.replace('\x00BS\x00', '\\')

    # Collapse multiple spaces
    text = re.sub(r' {2,}', ' ', text)
    return text


# ═════════════════════════════════════════════════════════════════════════════
# FORMAT-SPECIFIC PARSERS
# ═════════════════════════════════════════════════════════════════════════════

def load_reference_dates(script_dir, source_num):
    """Load dates from the reference worksheet for a given source number."""
    ref_path = script_dir / REFERENCE_WORKSHEET
    if not ref_path.exists():
        return []
    try:
        df = pd.read_excel(ref_path)
        src_df = df[df['Sources'] == source_num].reset_index(drop=True)
        dates = []
        for val in src_df['Date']:
            if pd.isna(val):
                dates.append(None)
            elif isinstance(val, datetime):
                dates.append(val)
            elif isinstance(val, str):
                # Try to parse string dates (handles typos like "21-0ct-04")
                fixed = val.replace('0ct', 'Oct').replace('0CT', 'OCT')
                try:
                    dates.append(pd.to_datetime(fixed).to_pydatetime())
                except Exception:
                    print(f"  WARNING: Could not parse date string: '{val}'")
                    dates.append(None)
            else:
                dates.append(None)
        return dates
    except Exception as e:
        print(f"  WARNING: Could not load reference dates: {e}")
        return []


def parse_government(content, source_name, source_num, ref_dates=None):
    """Government: asterisk separators, strip 'Article N' prefix from title.
    Dates come from the reference worksheet (articles have no inline dates)."""
    chunks = [c.strip() for c in SEP_ASTERISKS.split(content) if c.strip()]
    rows = []

    for idx, chunk in enumerate(chunks):
        # Use reference date by position if available
        if ref_dates and idx < len(ref_dates):
            date = ref_dates[idx]
        else:
            date = None
        year = date.year if date else None

        # Skip year filter when using reference dates (reference is authoritative)
        if not ref_dates and not in_year_range(year):
            continue

        lines = get_non_blank_lines(chunk)
        if not lines:
            continue

        # Skip "Article N" prefix line
        start = 0
        if re.match(r'^Article\s+\d+\s*$', lines[0], re.IGNORECASE):
            start = 1

        if start >= len(lines):
            continue

        title = lines[start]
        body_text = _body_after_title(chunk, title)
        body, refs = strip_references(body_text)
        rows.append(make_row(date, year, source_num, source_name, title, body, refs))

    return rows


def parse_after_label(content, source_name, source_num):
    """Science Research / Science News: section label precedes title."""
    chunks = [c.strip() for c in SEP_ASTERISKS.split(content) if c.strip()]
    rows, skipped = [], 0

    for chunk in chunks:
        date = extract_date(chunk)
        year = date.year if date else None
        if not in_year_range(year):
            skipped += 1
            continue

        lines = get_non_blank_lines(chunk)
        if not lines:
            continue

        # Find title: line after a section label
        title = ""
        for i, line in enumerate(lines):
            if SECTION_LABELS.match(line) and i + 1 < len(lines):
                candidate = lines[i + 1]
                if not SCIENCE_META.match(candidate) and not re.match(r'^\d+$', candidate):
                    title = candidate
                    break

        # Fallback: first substantive non-metadata line
        if not title:
            for line in lines:
                if (not SCIENCE_META.match(line)
                        and not re.match(r'^\d+$', line)
                        and len(line) > 3):
                    title = line
                    break

        if not title:
            continue

        # Body after title, strip leading science metadata
        body_text = _body_after_title(chunk, title)
        body_lines = body_text.splitlines()
        while body_lines and (not body_lines[0].strip()
                              or SCIENCE_META.match(body_lines[0].strip())):
            body_lines.pop(0)

        body_text = "\n".join(body_lines).strip()
        body, refs = strip_references(body_text)
        rows.append(make_row(date, year, source_num, source_name, title, body, refs))

    if skipped:
        print(f"  (skipped {skipped} articles outside {YEAR_MIN}–{YEAR_MAX})")
    return rows


def _strip_business_tail(lines):
    """Remove trailing junk (Document ID, 'More Like This', contact blocks)."""
    while lines:
        last = lines[-1].strip()
        if not last or BUSINESS_TAIL_JUNK.match(last):
            lines.pop()
        else:
            break
    # Strip trailing contact block (has both phone AND email on same line)
    if lines and (re.search(r'\d{3}[.-]\d{3}[.-]\d{4}', lines[-1])
                  and re.search(r'\b\S+@\S+\.\S+', lines[-1])):
        lines.pop()
    return lines


def parse_business(content, source_name, source_num):
    """Business Press / RTF: 'Article N ****' separator, real headline as title."""
    chunks = [c.strip() for c in SEP_BUSINESS.split(content) if c.strip()]
    rows, skipped = [], 0

    for chunk in chunks:
        date = extract_date(chunk)
        year = date.year if date else None
        if not in_year_range(year):
            skipped += 1
            continue

        lines = get_non_blank_lines(chunk)
        if not lines:
            continue

        title = lines[0]  # actual headline

        # Body starts after copyright line
        raw_lines = chunk.splitlines()
        body_start = 0
        for i, line in enumerate(raw_lines):
            if COPYRIGHT_LINE.match(line.strip()):
                body_start = i + 1
                break

        if body_start == 0:
            # Fallback: skip metadata block (~first 8 lines)
            body_start = min(8, len(raw_lines))

        body_lines = list(raw_lines[body_start:])
        body_lines = _strip_business_tail(body_lines)
        body_text = "\n".join(body_lines).strip()
        body, refs = strip_references(body_text)

        if count_words(body) < 5:
            continue

        rows.append(make_row(date, year, source_num, source_name, title, body, refs))

    if skipped:
        print(f"  (skipped {skipped} articles outside {YEAR_MIN}–{YEAR_MAX})")
    return rows


def parse_futurist(content, source_name, source_num):
    """Futurists: split by ToC lines and asterisk separators."""
    combined_sep = re.compile(
        r'(?:^\s*Foresight Update \d+\s*-\s*Table of Contents.*$|^\s*\*{5,}\s*$)',
        re.MULTILINE
    )
    chunks = [c.strip() for c in combined_sep.split(content) if c.strip()]

    # Lines that appear in issue headers (not article content)
    header_line_re = re.compile(
        r'^(A publication of the Foresight Institute|'
        r'Preparing for future technologies|'
        r'Board of Directors|'
        r'All Rights Reserved|'
        r'Write to the Foresight Institute|'
        r'If you find information|'
        r'Editor\s|Publisher\s|'
        r'.{0,30}(President|Secretary|Treasurer)\s*$|'
        r'Box \d+.*CA\s+\d|'
        r'.{0,3}Copyright\s+\d{4}|'
        r'Foresight Institute\s*$|'
        r'\d{1,2}\s+\w+\s+\d{4}\s*$)',
        re.IGNORECASE
    )

    # Whole chunks to skip entirely
    skip_chunk_re = re.compile(
        r'^(Clippings Invited|'
        r'If you find information and clippings|'
        r'Write to the Foresight Institute)',
        re.IGNORECASE
    )

    rows, skipped = [], 0
    last_date = None

    for chunk in chunks:
        lines = get_non_blank_lines(chunk)
        if not lines:
            continue

        # Always try to capture date (even from skipped header chunks)
        chunk_date = extract_date(chunk)
        if chunk_date:
            last_date = chunk_date

        # Skip known boilerplate chunks
        if any(skip_chunk_re.match(l) for l in lines[:2]):
            continue
        if count_words(chunk) < 20:
            continue

        # Strip issue-header boilerplate from chunk start (only when
        # the chunk actually begins with a known header line)
        if lines and header_line_re.match(lines[0]):
            while lines:
                if header_line_re.match(lines[0]):
                    lines.pop(0)
                elif len(lines[0].split()) <= 2:
                    lines.pop(0)  # short name lines in header block
                elif lines[0][0].islower():
                    lines.pop(0)  # continuation of previous header paragraph
                else:
                    break

        if not lines or count_words(' '.join(lines)) < 15:
            continue

        date = chunk_date if chunk_date else last_date
        year = date.year if date else None
        if not in_year_range(year):
            skipped += 1
            continue

        title = lines[0]
        body = _body_after_title(chunk, title)
        rows.append(make_row(date, year, source_num, source_name, title, body, ""))

    if skipped:
        print(f"  (skipped {skipped} articles outside {YEAR_MIN}–{YEAR_MAX})")
    return rows


# ═════════════════════════════════════════════════════════════════════════════
# DISPATCHER
# ═════════════════════════════════════════════════════════════════════════════

def parse_articles(filepath, source_name, source_num, title_style):
    """Route to the appropriate parser."""
    script_dir = filepath.parent

    if title_style == "business_rtf":
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            raw = f.read()
        content = strip_rtf(raw)
        content = content.replace('\r\n', '\n').replace('\r', '\n')
        return parse_business(content, source_name, source_num)

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()
    content = content.replace('\r\n', '\n').replace('\r', '\n')

    if title_style == "government":
        ref_dates = load_reference_dates(script_dir, source_num)
        if ref_dates:
            print(f"  (loaded {len(ref_dates)} dates from {REFERENCE_WORKSHEET})")
        return parse_government(content, source_name, source_num, ref_dates)
    elif title_style == "after_label":
        return parse_after_label(content, source_name, source_num)
    elif title_style == "business_press":
        return parse_business(content, source_name, source_num)
    elif title_style == "futurist":
        return parse_futurist(content, source_name, source_num)
    else:
        raise ValueError(f"Unknown title_style: {title_style}")


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═════════════════════════════════════════════════════════════════════════════

def write_excel(all_rows, output_path):
    columns = (
        ['Date', 'Year', 'Sources', 'Name', 'Word count', 'Title', 'Body', 'References']
        + list(TOPIC_KEYWORDS.keys())
    )
    df = pd.DataFrame(all_rows, columns=columns)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        ws = writer.sheets['Sheet1']

        header_fill = PatternFill('solid', start_color='4472C4', end_color='4472C4')
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        fixed_widths = {'A': 14, 'B': 6, 'C': 8, 'D': 18, 'E': 12}
        all_cols = [chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26)
                    for i in range(ws.max_column)]
        for col_idx, col_letter in enumerate(all_cols, start=1):
            col_name = df.columns[col_idx - 1]
            if col_name == 'Title':
                ws.column_dimensions[col_letter].width = 40
            elif col_name in ('Body', 'References'):
                ws.column_dimensions[col_letter].width = 60
            else:
                ws.column_dimensions[col_letter].width = fixed_widths.get(col_letter, 16)

        data_font = Font(name='Arial', size=11)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = Alignment(wrap_text=False, vertical='top')


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

def main():
    script_dir = Path(__file__).parent
    all_rows = []

    for filename, (source_name, source_num, title_style) in FILE_SOURCE_MAP.items():
        filepath = script_dir / filename
        if not filepath.exists():
            print(f"WARNING: '{filename}' not found in {script_dir} — skipping.")
            continue

        rows = parse_articles(filepath, source_name, source_num, title_style)
        all_rows.extend(rows)
        print(f"'{filename}' ({source_name}, Sources={source_num}): {len(rows)} articles")
        # Print a few samples for spot-checking
        for r in rows[:3]:
            safe_title = r['Title'][:60].encode('ascii', 'replace').decode('ascii')
            print(f"  Sample: Date={r['Date']}, Words={r['Word count']}, "
                  f"Title='{safe_title}'")

    if not all_rows:
        print("No articles found. Check that your source files are in the same folder.")
        return

    output_path = script_dir / OUTPUT_FILE
    write_excel(all_rows, output_path)
    print(f"\nDone! {len(all_rows)} total rows written to '{output_path}'")


if __name__ == '__main__':
    main()
